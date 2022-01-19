from matplotlib import pyplot #Импортировать библиотеки pyplot
from openpyxl import load_workbook # Импортировать библиотеки load_workbook

wb = load_workbook('data_analysis_lab.xlsx')
sheet = wb ['Data']

def getvalue(x):
    return x.value
