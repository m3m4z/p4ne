from matplotlib import pyplot #Импортировать библиотеки pyplot
from openpyxl import load_workbook #Импортировать библиотеки load_workbook

wb = load_workbook('data_analysis_lab.xlsx') #загрузка таблицы с данными
sheet = wb ['Data']

def getvalue(x):
    return x.value

years = list(map(getvalue, sheet['A'][1:]))
temperature = list(map(getvalue, sheet['C'][1:]))
activity = list(map(getvalue, sheet['D'][1:]))

#График
pyplot.plot(years, temperature, label="Относит. Темпрература")
pyplot.plot(years, activity, Label="Активность Солнца")

#Показ графика
pyplot.xlabel('Годы')
pyplot.ylabel('Температура\Активность солнца')
pyplot.show()
